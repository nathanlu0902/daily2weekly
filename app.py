﻿
from openpyxl import Workbook, load_workbook
from collections import defaultdict
from openpyxl.styles import PatternFill, Color, Border, Side
from openpyxl.utils import column_index_from_string


class part_intersection:

    def __init__(self, pn, vendor_code, site, hub, onway, vendor_name):
        self.PN = pn
        self.vendorCode = vendor_code
        self.site = site
        self.hub = hub
        self.onway = onway
        self.vendor_name = vendor_name

        self._commit = defaultdict()
        self._request = defaultdict()
        self._delta = defaultdict()

    def load_request(self, date_list, qty_list):

        if len(date_list) == len(qty_list):
            self.request = dict(zip(date_list, qty_list))

            for k, v in self.request.items():
                if v is None:
                    self.request[k] = 0

    def load_commit(self, date_list, qty_list):

        if len(date_list) == len(qty_list):
            self.commit = dict(zip(date_list, qty_list))
            for k, v in self.commit.items():
                if v is None:
                    self.commit[k] = 0

    def load_delta(self, date_list, qty_list):

        if len(date_list) == len(qty_list):
            self.delta = dict(zip(date_list, qty_list))
            for k, v in self.delta.items():
                if v is None:
                    self.delta[k] = 0


    def get_aggregate_request(self, cut_off_day, date_dic):

        i = 0
        self.request_dic_aggr = defaultdict()
        key = ''

        for date in self.request:
            if date == list(self.request.keys())[
                    0] or date_dic[date] == cut_off_day:
                key = date
                self.request_dic_aggr[key] = self.request[key]
                i += 1

            else:
                self.request_dic_aggr[key] += self.request[date]

        return self.request_dic_aggr

    def get_aggregate_commit(self, cut_off_day, date_dic):

        i = 0
        self.commit_dic_aggr = defaultdict()
        key = ''

        for date in self.commit:
            if date == list(self.commit.keys())[
                    0] or date_dic[date] == cut_off_day:
                key = date
                self.commit_dic_aggr[key] = self.commit[key]
                i += 1

            else:
                self.commit_dic_aggr[key] += self.commit[date]

    def get_delta(self, cut_off_day, date_dic):

        i = 0
        self.delta_dic = defaultdict()

        for date in self.delta:
            if date_dic[date] == cut_off_day:
                self.delta_dic[date] = self.delta[date]
                i += 1


    @property
    def weekly_commit(self):
        return self.commit_dic_aggr

    @property
    def weekly_request(self):
        return self.request_dic_aggr


class Excel_handler:

    def __init__(self, path, cut_off_day):

        # #open the Excel and save it so that value can be calculated
        # with open(path,'a') as f:
        #     f.write('2')
        #     f.close()

        self.wb=load_workbook(path,data_only=True)

        self.ws = self.wb.active
        self.part_count = self.part_counts()
        self.date_dic = self.get_calendar_dict()
        self.CutOffDay = cut_off_day

    # def test_save(self):
    #     self.wb.save('test.xlsx')

    def part_counts(self):
        # count how many parts in workbook
        row_count = self.ws.max_row
        part_count = round(row_count / 15)
        return part_count

    def get_calendar_dict(self):
        date_dic = defaultdict()

        for col in self.ws.iter_cols(
                min_col=column_index_from_string('O'),
                min_row=2,
                max_row=3,
                values_only=True):
            date = col[0]
            weekday = col[1]
            date_dic[date] = weekday

        return date_dic

    # load master data and sp data for each intersection

    def load_data(self):

        self.part_list = []

        for i in range(self.part_count):
            PN = self.ws.cell(row=4 + 15 * i, column=4).value
            vendorCode = self.ws.cell(row=4 + 15 * i, column=1).value
            site = self.ws.cell(row=4 + 15 * i, column=3).value
            hub = self.ws.cell(
                row=4 + 15 * i,
                column=column_index_from_string("M")).value
            onway = self.ws.cell(
                row=4 + 15 * i,
                column=column_index_from_string('N')).value
            vendor_name = self.ws.cell(row=4 + 15 * i, column=2).value

            request_list = []
            commit_list = []
            delta_list = []

            for row in self.ws.iter_rows(
                    min_col=column_index_from_string('O'),
                    min_row=4 + 15 * i,
                    max_row=4 + 15 * i):
                for cell in row:
                    request_list.append(cell.value)

            for row in self.ws.iter_rows(
                    min_col=column_index_from_string('O'),
                    min_row=6 + 15 * i,
                    max_row=6 + 15 * i):
                for cell in row:
                    commit_list.append(cell.value)

            for row in self.ws.iter_rows(
                    min_col=column_index_from_string('O'),
                    min_row=9 + 15 * i,
                    max_row=9 + 15 * i):
                for cell in row:
                    delta_list.append(cell.value)

            part = part_intersection(
                PN, vendorCode, site, hub, onway, vendor_name)

            part.load_request(self.date_dic.keys(), request_list)
            part.load_commit(self.date_dic.keys(), commit_list)
            part.load_delta(self.date_dic.keys(), delta_list)

            part.get_aggregate_request(self.CutOffDay, self.date_dic)
            part.get_aggregate_commit(self.CutOffDay, self.date_dic)
            part.get_delta(self.CutOffDay, self.date_dic)

            self.part_list.append(part)

    def write_to_excel(self, output):

        self.load_data()

        if len(self.part_list) == 0:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = 'Weekly Supply'

        # write the first row
        first_row_data = [
            'APN',
            'Apple Vendor Code',
            'Vendor Name',
            'Site Code',
            'Data Type',
            'VMI Hub',
            'VMI Onway']
        # calendar
        first_row_data.extend(list(self.part_list[0].weekly_request.keys()))
        week_days = [self.date_dic[day]
                     for day in self.part_list[0].weekly_request.keys()]
        ws.append(first_row_data)

        second_row_data = [
            '',
            '',
            '',
            '',
            '',
            '',
            '']
        second_row_data.extend(week_days)
        ws.append(second_row_data)

        data_list = []

        for i in range(0, len(self.part_list)):
            part = self.part_list[i]
            req_data = [
                part.PN,
                part.vendorCode,
                part.vendor_name,
                part.site,
                'Request',
                '',
                '']
            req_data.extend(list(part.weekly_request.values()))
            data_list.append(req_data)

            commit_data = [
                part.PN,
                part.vendorCode,
                part.vendor_name,
                part.site,
                'Commit',
                part.hub,
                part.onway]
            commit_data.extend((list(part.weekly_commit.values())))
            data_list.append(commit_data)

            delta_data = [
                part.PN,
                part.vendorCode,
                part.vendor_name,
                part.site,
                'Cum Delta',
                '',
                '']
            delta_data.extend((list(part.delta_dic.values())))
            data_list.append(delta_data)

        for j in range(0, len(data_list), 2):
            ws.append(data_list[j])
            ws.append(data_list[j + 1])

        # define boarder
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        # fill in background color
        for n in range(3, ws.max_row + 1):
            cells = ws[n:n]
            if n % 3 == 0:
                for cell in cells:
                    cell.fill = PatternFill(
                        fgColor=Color('CCCCFF'), fill_type='solid')
            elif (n - 1) % 3 == 0:
                for cell in cells:
                    cell.fill = PatternFill(
                        fgColor=Color('FFFFCC'), fill_type='solid')
            else:
                for cell in cells:
                    cell.fill = PatternFill(
                        fgColor=Color('5DADD5'), fill_type='solid')

        wb.save(output)
